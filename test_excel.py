
import os
import csv
from openpyxl import Workbook

# Create reports directory if it doesn't exist
reports_dir = 'reports'
os.makedirs(reports_dir, exist_ok=True)

# Sample data for testing
broken_links_data = [
    ['Source Page URL', 'Broken Link URL', 'Anchor Text / Current Value', 'Link Type', 'Status Code'],
    ['https://hosninsurance.ae/', 'https://hosninsurance.ae/old-services-page', 'Our Services (Outdated)', 'Internal', '404'],
    ['https://hosninsurance.ae/about', 'https://facebook.com/company-old-page', 'Follow us on Facebook', 'External', '404'],
    ['https://hosninsurance.ae/contact', 'https://hosninsurance.ae/resources/company-brochure.pdf', 'Download Company Brochure', 'Internal', '404'],
    ['https://hosninsurance.ae/services', 'https://twitter.com/company_handle_old', 'Twitter Updates', 'External', '404'],
    ['https://hosninsurance.ae/', 'https://hosninsurance.ae/news/press-release-2023', 'Latest Press Release', 'Internal', '404']
]

orphan_pages_data = [
    ['Orphan Page URL', 'Found in Sitemap?', 'Internally Linked?'],
    ['https://hosninsurance.ae/legacy/old-product-page', 'Yes', 'No'],
    ['https://hosninsurance.ae/archived/company-history', 'Yes', 'No'],
    ['https://hosninsurance.ae/temp/beta-features', 'Yes', 'No'],
    ['https://hosninsurance.ae/old-blog/category/updates', 'Yes', 'No'],
    ['https://hosninsurance.ae/hidden/internal-tools', 'Yes', 'No']
]

referring_domains_data = [
    ['Domain', 'Domain Rating', 'Spam Score', 'Backlinks', 'Link Type', 'First Seen', 'Target Page', 'Anchor Text'],
    ['google.com', '100', '0%', '45', 'DoFollow', '2024-01-15', 'Homepage', 'Brand name'],
    ['facebook.com', '96', '2%', '23', 'NoFollow', '2024-02-10', 'About page', 'Company profile'],
    ['linkedin.com', '95', '1%', '34', 'DoFollow', '2024-01-20', 'Homepage', 'Professional services'],
    ['twitter.com', '94', '3%', '18', 'NoFollow', '2024-03-05', 'Blog posts', 'Social mention'],
    ['wikipedia.org', '93', '0%', '12', 'DoFollow', '2024-01-28', 'References', 'Citation link']
]

try:
    # Create workbook
    wb = Workbook()
    
    # Create Excel filename
    excel_filename = "report_hosninsurance_ae.xlsx"
    excel_filepath = os.path.join(reports_dir, excel_filename)
    
    # 1. Broken Links Sheet
    ws1 = wb.active
    ws1.title = "Broken"
    for row in broken_links_data:
        ws1.append(row)
    
    # 2. Orphan Pages Sheet
    ws2 = wb.create_sheet("Orphan")
    for row in orphan_pages_data:
        ws2.append(row)
    
    # 3. Referring Domains Sheet
    ws3 = wb.create_sheet("Referring")
    for row in referring_domains_data:
        ws3.append(row)
    
    # Save Excel file
    wb.save(excel_filepath)
    print(f"Successfully generated Excel report: {excel_filename}")
    print(f"File saved to: {excel_filepath}")
    print(f"File size: {os.path.getsize(excel_filepath)} bytes")
    
    # List all files in reports directory
    print(f"\nFiles in reports directory:")
    for file in os.listdir(reports_dir):
        filepath = os.path.join(reports_dir, file)
        size = os.path.getsize(filepath)
        print(f"  - {file} ({size} bytes)")
        
except Exception as e:
    print(f"Error generating Excel file: {e}")
